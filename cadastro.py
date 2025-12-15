import pandas as pd
import streamlit as st
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import re

arquivo_excel = Path("cadastro.xlsx")

colunas = [
    "Nº Reclamação", "VR", "Data Recebimento", "Nome/Razão Social", "Telefone", "Email",
    "Processo Relacionado a Reclamação", "Canal de Entrada da Reclamação",
    "Descrição da Reclamação", "Procedente/Não Procedente", "Forma de Retorno",
    "Descrição da Resposta", "Status do Retorno ao Cliente", "Data do Retorno",
    "Ação tomada", "Custo da Ação"
]

if not arquivo_excel.exists():
    df = pd.DataFrame(columns=colunas)
    df.to_excel(arquivo_excel, index=False)

def validar_email(email):
    return re.match(r"[^@]+@[^@]+\.[^@]+", email)

def validar_telefone(telefone):
    return telefone.isdigit() and (8 <= len(telefone) <= 15)

if 'refresh' not in st.session_state:
    st.session_state['refresh'] = 0

st.markdown("""
<style>
.block-container { padding-top: 2rem !important; padding-bottom: 2rem !important; }
div.stButton > button {
    display: block; margin-left: auto; margin-right: auto;
    background-color: #4CAF50; color: white; height: 50px; width: 220px;
    border-radius: 10px; font-size: 18px; margin-top: 15px;
}
div[data-baseweb="input"] > div > input, textarea, select {
    border: 2px solid #4CAF50; border-radius: 5px; padding: 8px;
}
body { background-color: #f0f2f6; }
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 style='text-align:center; color:#4CAF50;'>Cadastro de Reclamação</h1>", unsafe_allow_html=True)

col1, col2, col3 = st.columns(3)

with col1:
    vr = st.text_input("VR")
    nome = st.text_input("Nome/Razão Social")
    telefone = st.text_input("Telefone")
    processo = st.selectbox("Processo Relacionado a Reclamação", ["Atendimento", "Bilheteria", "Comercial", "Financeiro", "Marketing", "Operacional", "Parque Temático", "Maria Fumaça", "Produto e Logística", "Qualidade", "Receptivo", "Reservas", "Site / Serviços Digitais", "Sugestão / Outros"])
    procedente = st.selectbox("Procedente/Não Procedente", ["Procedente", "Não Procedente"])
    
with col2:
    data_recebimento = st.date_input("Data Recebimento")
    email = st.text_input("Email")
    canal = st.selectbox("Canal de Entrada da Reclamação", ["Atendimento", "Atraso / Horários", "Venda / Ingressos", "Pagamento / Reembolso", "Site / Sistema", "Acessibilidade", "Passeio / Transporte", "Atrações / Espetáculos", "Infraestrutura / Conforto", "Outros"]
)
    forma_retorno = st.selectbox("Forma de Retorno", ["Telefone", "WhatsApp", "E-mail", "Presencial", "Redes Sociais", "Reclame Aqui", "Google / Avaliações", "Pesquisa de Satisfação", "Procon / Sistema Oficial", "Outros"])
    status_retorno = st.selectbox("Status do Retorno ao Cliente", ["Concluído", "Não Concluído", "Em Andamento"])

with col3:
    descricao_reclamacao = st.text_input("Descrição da Reclamação")
    descricao_resposta = st.text_input("Descrição da Resposta")
    data_retorno = st.date_input("Data do Retorno")
    acao_tomada = st.text_input("Ação tomada")
    custo_acao = st.text_input("Custo da Ação")

def formatar_excel(arquivo):
    wb = load_workbook(arquivo)
    ws = wb.active

    larguras = {
        "A": 15, "B": 10, "C": 15, "D": 25, "E": 15,
        "F": 25, "G": 30, "H": 25, "I": 40, "J": 20,
        "K": 20, "L": 40, "M": 25, "N": 15, "O": 40, "P": 15
    }
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = 'DD/MM/YYYY'
    for row in ws.iter_rows(min_row=2, min_col=13, max_col=13):
        for cell in row:
            cell.number_format = 'DD/MM/YYYY'

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

    for table in ws._tables:
        ws._tables.remove(table)

    tab = Table(displayName="TabelaCadastro", ref=f"A1:P{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(arquivo)

if st.button("Salvar"):

    if not nome or not email or not telefone:
        st.error("Preencha os campos obrigatórios: Nome/Razão Social, Email e Telefone.")
    elif not validar_email(email):
        st.error("E-mail inválido.")
    elif not validar_telefone(telefone):
        st.error("Telefone inválido. Apenas números, entre 8 e 15 dígitos.")
    else:
        df = pd.read_excel(arquivo_excel)

        if df.empty:
            numero_reclamacao = 1
        else:
            numero_reclamacao = df["Nº Reclamação"].max() + 1

        novo_cadastro = [[
            numero_reclamacao, vr, data_recebimento, nome, telefone, email,
            processo, canal, descricao_reclamacao, procedente,
            forma_retorno, descricao_resposta, status_retorno,
            data_retorno, acao_tomada, custo_acao
        ]]

        df = pd.concat(
            [df, pd.DataFrame(novo_cadastro, columns=df.columns)],
            ignore_index=True
        )

        df.to_excel(arquivo_excel, index=False)

        formatar_excel(arquivo_excel)

        st.success(f"Cadastro Nº {numero_reclamacao} salvo com sucesso!")
        st.session_state['refresh'] += 1

st.subheader("Cadastros existentes")
df = pd.read_excel(arquivo_excel)
st.dataframe(df, height=400)
